# ===========================================
#                Note to New Users
# ===========================================

# Please ensure you have completed the following prerequisite steps to avoid errors:

# 1. Install the required modules:
#    - Use the modules listed in 'Python-Module-Requirement.txt' or as shown at the beginning of the code.

# 2. Ensure all dependency files and folders are within the same directory as this script.

# Dependency File List:
# ---------------------
# - CDR_Processed_Industrial_Processing_and_Use.xlsx
# - CDR_Processed_Consumer_and_Commercial_Use.xlsx
# - TRI_US_1b_2020.xlsx
# - Directory: 'ChemExpo Bulk Composition XLSX' containing files such as:
#     - ChemExpo_bulk_composition_chemicals-1.xlsx
# - PUCs_NAICS Crosswalk_ERG Draft_06-19-2024_UPDATED.xlsx
# - 2022-NAICS-Codes-listed-numerically-2-Digit-through-6-Digit.xlsx



import pandas as pd
import re
import os
import time
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from reportlab.lib.pagesizes import portrait, letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from diagrams import Diagram, Cluster, Edge
from diagrams.custom import Custom

# Helper function to convert CAS number to 10-digit format
def format_cas_number(cas):
    parts = cas.split('-')
    return f"{int(parts[0]):06}{int(parts[1]):02}{int(parts[2]):01}"

# Helper function to convert CASRN to 10-digit number
def convert_casrn_to_10_digit(casrn):
    stripped_casrn = casrn.replace("-", "")
    return stripped_casrn.zfill(10)

# Helper function to extract CAS number from file name
def extract_cas_number(file_name, pattern):
    match = re.search(pattern, file_name)
    return match.group(1) if match else 'Unknown-CAS'

# Step 1: Process CDR and TRI data
def process_cdr_tri_data(casrn):
    # Read the TRI, CDR industrial, and CDR consumer data
    tri_df = pd.read_excel("TRI_US_1b_2020.xlsx")
    cdr_industrial_df = pd.read_excel('CDR_Processed_Industrial_Processing_and_Use.xlsx')
    cdr_consumer_df = pd.read_excel('CDR_Processed_Consumer_and_Commercial_Use.xlsx')

    # Concatenate CDR industrial and consumer data into one dataframe
    cdr_combined_df = pd.concat([cdr_industrial_df, cdr_consumer_df], ignore_index=True)

    # Convert casrn to string
    casrn = str(casrn)

    # Filter TRI data based on CASRN
    filtered_tri = tri_df[tri_df['76. CAS NUMBER'] == casrn]

    if not filtered_tri.empty:
        print("Chemical information found in TRI:")
        print(filtered_tri)

        # Filter combined CDR data based on CHEMICAL ID
        filtered_cdr_combined_df = cdr_combined_df[cdr_combined_df['CHEMICAL ID'] == casrn]

        if not filtered_cdr_combined_df.empty:
            print("Matching CDR data:")
            print(filtered_cdr_combined_df)

            # Merge TRI and filtered CDR data based on FRS FACILITY ID and EPA FACILITY REGISTRY ID, keeping all TRI rows
            merged_data = pd.merge(filtered_tri, filtered_cdr_combined_df, how='left', left_on='74. FRS FACILITY ID', right_on='EPA FACILITY REGISTRY ID')

            # Rename columns for better readability based on the "Use Type"
            if 'INDUSTRIAL TYPE OF PROCESS OR USE OPERATION' in merged_data.columns:
                merged_data = merged_data.rename(columns={
                    'INDUSTRIAL FUNCTION CATEGORY': 'Industrial Function Category',
                    'IND FUNCT CAT OTHER DESC': 'Industrial Other Function Category',
                    'INDUSTRIAL TYPE OF PROCESS OR USE OPERATION': 'Industrial Process or Use Operation',
                    'PHYSICAL FORM(S)': 'Physical Form',
                    'EPA FACILITY REGISTRY ID': 'EPA FACILITY REGISTRY ID',
                    'SITE NAICS CODE 1': 'NAICS Code 1',
                    'SITE NAICS ACTIVITY 1': 'NAICS Activity 1',
                    'SITE NAICS CODE 2': 'NAICS Code 2',
                    'SITE NAICS ACTIVITY 2': 'NAICS Activity 2',
                    'SITE NAICS CODE 3': 'NAICS Code 3',
                    'SITE NAICS ACTIVITY 3': 'NAICS Activity 3'
                })
                # Standardize "Intermediate" and "Intermediates"
                if 'Industrial Function Category' in merged_data.columns:
                    merged_data['Industrial Function Category'] = merged_data['Industrial Function Category'].replace('Intermediate', 'Intermediates')

            if 'CONSUMER / COMMERCIAL PRODUCT CATEGORY' in merged_data.columns:
                merged_data = merged_data.rename(columns={
                    'CONS / COMM FUNCTION CATEGORY': 'C/C Function Category',
                    'CONS COMM FUNCT CAT OTHER DESC': 'C/C Other Function Category',
                    'CONSUMER / COMMERCIAL PRODUCT CATEGORY': 'C/C Product Category',
                    'PHYSICAL FORM': 'Physical Form',
                    'EPA FACILITY REGISTRY ID': 'EPA FACILITY REGISTRY ID',
                    'SITE NAICS CODE 1': 'NAICS Code 1',
                    'SITE NAICS ACTIVITY 1': 'NAICS Activity 1',
                    'SITE NAICS CODE 2': 'NAICS Code 2',
                    'SITE NAICS ACTIVITY 2': 'NAICS Activity 2',
                    'SITE NAICS CODE 3': 'NAICS Code 3',
                    'SITE NAICS ACTIVITY 3': 'NAICS Activity 3'
                })
                # Standardize "Intermediate" and "Intermediates"
                if 'C/C Function Category' in merged_data.columns:
                    merged_data['C/C Function Category'] = merged_data['C/C Function Category'].replace('Intermediate', 'Intermediates')

            # Check if '41. PRIMARY NAICS CODE' exists before renaming
            if '41. PRIMARY NAICS CODE' in merged_data.columns:
                merged_data = merged_data.rename(columns={'41. PRIMARY NAICS CODE': '41. NAICS Code 1'})

            # Add "Use Type" column based on whether it's industrial or consumer data
            if 'Industrial Process or Use Operation' in merged_data.columns:
                merged_data['Use Type'] = 'Industrial'
            elif 'C/C Product Category' in merged_data.columns:
                merged_data['Use Type'] = 'Consumer and Commercial'
            else:
                merged_data['Use Type'] = ''

            # Drop duplicate rows
            merged_data.drop_duplicates(inplace=True)

            # Save merged data to Excel file
            output_file_name = f"TRI_CDR_REVERSED_{casrn}.xlsx"
            merged_data.to_excel(output_file_name, index=False)
            print(f"Merged CDR data has been saved to '{output_file_name}'.")
            return output_file_name
        else:
            print("No matching CDR data found.")

            # Save empty TRI data to Excel file
            empty_tri_df = filtered_tri.copy()  # Create an empty DataFrame with the same structure as filtered_tri
            output_file_name = f"TRI_REVERSED_{casrn}.xlsx"
            empty_tri_df.to_excel(output_file_name, index=False)
            print(f"Empty TRI data has been saved to '{output_file_name}'.")
            return output_file_name
    else:
        print("CAS number input is not registered within TRI. Attempting to search through CDR Database for matches.")

        # Create empty columns for TRI data
        empty_tri_df = pd.DataFrame(columns=tri_df.columns)

        # Filter combined CDR data based on CHEMICAL ID
        filtered_cdr_combined_df = cdr_combined_df[cdr_combined_df['CHEMICAL ID'] == casrn]

        if not filtered_cdr_combined_df.empty:
            print("Matching CDR data:")
            print(filtered_cdr_combined_df)

            # Combine empty TRI data with CDR data without merging
            combined_data = pd.concat([empty_tri_df, filtered_cdr_combined_df], axis=1)

            # Rename columns for better readability based on the "Use Type"
            if 'INDUSTRIAL TYPE OF PROCESS OR USE OPERATION' in combined_data.columns:
                combined_data = combined_data.rename(columns={
                    'INDUSTRIAL FUNCTION CATEGORY': 'Industrial Function Category',
                    'IND FUNCT CAT OTHER DESC': 'Industrial Other Function Category',
                    'INDUSTRIAL TYPE OF PROCESS OR USE OPERATION': 'Industrial Process or Use Operation',
                    'PHYSICAL FORM(S)': 'Physical Form',
                    'EPA FACILITY REGISTRY ID': 'EPA FACILITY REGISTRY ID',
                    'SITE NAICS CODE 1': 'NAICS Code 1',
                    'SITE NAICS ACTIVITY 1': 'NAICS Activity 1',
                    'SITE NAICS CODE 2': 'NAICS Code 2',
                    'SITE NAICS ACTIVITY 2': 'NAICS Activity 2',
                    'SITE NAICS CODE 3': 'NAICS Code 3',
                    'SITE NAICS ACTIVITY 3': 'NAICS Activity 3'
                })
                # Standardize "Intermediate" and "Intermediates"
                if 'Industrial Function Category' in combined_data.columns:
                    combined_data['Industrial Function Category'] = combined_data['Industrial Function Category'].replace('Intermediate', 'Intermediates')
            if 'CONSUMER / COMMERCIAL PRODUCT CATEGORY' in combined_data.columns:
                combined_data = combined_data.rename(columns={
                    'CONS / COMM FUNCTION CATEGORY': 'C/C Function Category',
                    'CONS COMM FUNCT CAT OTHER DESC': 'C/C Other Function Category',
                    'CONSUMER / COMMERCIAL PRODUCT CATEGORY': 'C/C Product Category',
                    'PHYSICAL FORM': 'Physical Form',
                    'EPA FACILITY REGISTRY ID': 'EPA FACILITY REGISTRY ID',
                    'SITE NAICS CODE 1': 'NAICS Code 1',
                    'SITE NAICS ACTIVITY 1': 'NAICS Activity 1',
                    'SITE NAICS CODE 2': 'NAICS Code 2',
                    'SITE NAICS ACTIVITY 2': 'NAICS Activity 2',
                    'SITE NAICS CODE 3': 'NAICS Code 3',
                    'SITE NAICS ACTIVITY 3': 'NAICS Activity 3'
                })
                # Standardize "Intermediate" and "Intermediates"
                if 'C/C Function Category' in combined_data.columns:
                    combined_data['C/C Function Category'] = combined_data['C/C Function Category'].replace('Intermediate', 'Intermediates')

            # Add "Use Type" column based on whether it's industrial or consumer data
            if 'Industrial Process or Use Operation' in combined_data.columns:
                combined_data['Use Type'] = 'Industrial'
            elif 'C/C Product Category' in combined_data.columns:
                combined_data['Use Type'] = 'Consumer and Commercial'
            else:
                combined_data['Use Type'] = ''

            # Save combined data to Excel file with the original naming convention
            output_file_name = f"TRI_CDR_REVERSED_{casrn}.xlsx"
            combined_data.to_excel(output_file_name, index=False)
            print(f"Combined CDR data has been saved to '{output_file_name}'.")
            return output_file_name
        else:
            print("No matching CDR data found.")

            # Save empty TRI data to Excel file
            output_file_name = f"TRI_REVERSED_{casrn}.xlsx"
            empty_tri_df.to_excel(output_file_name, index=False)
            print(f"Empty TRI data has been saved to '{output_file_name}'.")
            return output_file_name


# Step 2: Process ChemExpo data
def process_chemexpo_data(user_input):
    bulk_composition_directory = "ChemExpo Bulk Composition XLSX"
    data = []

    for filename in os.listdir(bulk_composition_directory):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(bulk_composition_directory, filename)
            df = pd.read_excel(file_path)
            filtered_df = df[df['Raw CAS'] == user_input]
            print(f"Scanning for chemical product information in '{filename}'")

            if not filtered_df.empty:
                print(f"Found matching chemical product information in '{filename}'")
                chemical_info = filtered_df.to_dict(orient='records')
                data.extend(chemical_info)

    if data:
        output_df = pd.DataFrame(data)
        output_file_name = f"Product_Use_Information_REVERSED_{user_input}.xlsx"
        output_df.to_excel(output_file_name, index=False)
        print(f"Tabulated chemical information has been saved to '{output_file_name}'.")
        return output_file_name
    else:
        print("No matching chemical information found.")
        return None

# Step 2.5: Match with PUCS NAICS Crosswalk
def match_with_pucs_naics(product_use_info_file, user_input):
    puc_naics_crosswalk_file = 'PUCs_NAICS Crosswalk_ERG Draft_06-19-2024_UPDATED.xlsx'
    puc_naics_df = pd.read_excel(puc_naics_crosswalk_file, sheet_name='FINAL RESULTS')
    product_use_df = pd.read_excel(product_use_info_file, sheet_name=0)

    # Rename the columns in puc_naics_df to match those in product_use_df
    puc_naics_df.rename(columns={
        'PUC_level1': 'PUC General Category',
        'PUC_level2': 'PUC Product Family',
        'PUC_level3': 'PUC Product Type'
    }, inplace=True)

    # Replace "nan" values with an empty string in the key columns
    match_columns = ['PUC General Category', 'PUC Product Family', 'PUC Product Type']
    for col in match_columns:
        product_use_df[col] = product_use_df[col].astype(str).replace('nan', '').str.strip().str.lower()
        puc_naics_df[col] = puc_naics_df[col].astype(str).replace('nan', '').str.strip().str.lower()

    # Merge the dataframes on the key columns
    merged_df_main = pd.merge(
        product_use_df, 
        puc_naics_df, 
        on=match_columns, 
        how='left')

    # Identify rows that need special condition handling
    special_condition_df = product_use_df[product_use_df['PUC Product Type'].eq('')]

    # Perform the merge for the special condition
    special_condition_merged_df = pd.merge(
        special_condition_df,
        puc_naics_df,
        how='left',
        left_on=['PUC General Category', 'PUC Product Family', 'PUC Product Family'],
        right_on=['PUC General Category', 'PUC Product Family', 'PUC Product Type']
    )
    
    # Combine results
    combined_df = pd.concat([merged_df_main, special_condition_merged_df])

    # Ensure that unmatched rows do not have any values in the NAICS columns
    naics_columns = ['NAICS 1', 'NAICS 2', 'NAICS 3']
    for col in naics_columns:
        if col in combined_df.columns:
            combined_df[col] = combined_df[col].where(pd.notnull(combined_df[col]), None)

    # Remove rows where all NAICS columns are None
    combined_df = combined_df[combined_df[naics_columns].notna().any(axis=1)]

    # Drop duplicates
    combined_df.drop_duplicates(inplace=True)

    # Save the merged data to a new Excel file
    output_file = f'Product_Use_Information_NAICS_Crosswalked_{user_input}.xlsx'
    combined_df.to_excel(output_file, index=False)
    print(f"Matched data with PUCS NAICS Crosswalk has been saved to '{output_file}'.")

    return output_file

# Step 3: Final matching and integration
def final_matching_and_integration(condition_of_use_file, pucs_naics_file, output_file):
    condition_of_use_df = pd.read_excel(condition_of_use_file)
    pucs_naics_df = pd.read_excel(pucs_naics_file)

    # Extract the CAS number from the input file names
    cas_number_condition = extract_cas_number(condition_of_use_file, r'TRI_CDR_REVERSED_(.+).xlsx')
    cas_number_pucs = extract_cas_number(pucs_naics_file, r'Product_Use_Information_NAICS_Crosswalked_(.+)')

    # Define possible NAICS columns
    naics_columns_condition_normal = [
        '41. NAICS CODE 1', '42. NAICS CODE 2', '43. NAICS CODE 3', 
        '44. NAICS CODE 4', '45. NAICS CODE 5', '46. NAICS CODE 6'
    ]
    naics_columns_condition_special = ['NAICS Code 1', 'NAICS Code 2', 'NAICS Code 3']
    naics_columns_pucs = ['NAICS 1', 'NAICS 2', 'NAICS 3']
    description_columns_pucs = ['NAICS 1 Description', 'NAICS 2 Description', 'NAICS 3 Description']
    required_columns = [
        "Data Document Title", "Data Document Subtitle", "Document Date", "Product Name", 
        "PUC Kind", "PUC General Category", "PUC Product Family", "PUC Product Type", "Raw Min Comp", 
        "Raw Max Comp", "Raw Central Comp", "Unit Type", "Lower Weight Fraction", 
        "Upper Weight Fraction", "Central Weight Fraction", "Weight Fraction Type", 
        "Component", "PUCID"
    ]

    # Determine which set of NAICS columns to use
    if set(naics_columns_condition_special).issubset(condition_of_use_df.columns):
        naics_columns_condition = naics_columns_condition_special
    else:
        naics_columns_condition = naics_columns_condition_normal

    # Prepare the result DataFrame with additional columns for matches
    result_df = condition_of_use_df.copy()
    for i in range(3):
        result_df[f'Matched NAICS {i+1}'] = ''
        result_df[f'Matched NAICS {i+1} Description'] = ''
        for col in required_columns:
            result_df[f'Matched NAICS {i+1} - {col}'] = ''

    # Iterate over each row in the condition_of_use_df
    for index, row in condition_of_use_df.iterrows():
        for i, naics_col in enumerate(naics_columns_condition):
            if naics_col in row and pd.notna(row[naics_col]):
                naics_code = str(row[naics_col])[:6]  # Truncate to first 6 characters
                if naics_code.isdigit():  # Ensure the NAICS code is numeric
                    # Check each NAICS column in PUCS for matches
                    for j in range(3):
                        matching_rows = pucs_naics_df[
                            (pucs_naics_df[naics_columns_pucs[j]].astype(str) == naics_code)
                        ]

                        if not matching_rows.empty:
                            for _, match_row in matching_rows.iterrows():
                                result_df.at[index, f'Matched NAICS {i+1}'] = match_row[naics_columns_pucs[j]]
                                result_df.at[index, f'Matched NAICS {i+1} Description'] = match_row[description_columns_pucs[j]]
                                for col in required_columns:
                                    result_df.at[index, f'Matched NAICS {i+1} - {col}'] = match_row[col]

    # Add "Use Type" column based on whether it's industrial or consumer data
    result_df['Use Type'] = result_df.apply(lambda row: 'Industrial' if pd.notnull(row['Industrial Process or Use Operation']) or pd.notnull(row['Industrial Function Category']) else ('Consumer and Commercial' if pd.notnull(row['C/C Product Category']) or pd.notnull(row['C/C Function Category']) else ''), axis=1)

    # Save the final merged data to a new Excel file
    result_df.to_excel(output_file, index=False)

    # Check the resulting dataframe
    print(result_df.shape)
    print(result_df.head())    
   
    print(f"Final merged data has been saved to '{output_file}'.")

    return output_file


# Step 4 - Data Processing before generating a facility summary report and qualitative flow diagram
def generate_qualitative_summary(final_merged_file, cas_number):
    # Load the Excel files (Change the file as you see fit)
    file_path = final_merged_file
    naics_file_path = '2022-NAICS-Codes-listed-numerically-2-Digit-through-6-Digit.xlsx'

    # Read the data from the files above
    df = pd.read_excel(file_path, sheet_name=0)
    naics_df = pd.read_excel(naics_file_path, sheet_name='Six Digit NAICS')

    # Extract CAS number from file path
    cas_number_match = re.search(r'Condition-of-use-TRI1b-CDR-PUCS-NAICS_(\d{2,7}-\d{2}-\d)', final_merged_file)
    if cas_number_match:
        cas_number = cas_number_match.group(1)
    else:
        raise ValueError("CAS number not found in the file path.")

    # Extracting relevant columns
    activity_col = 'ACTIVITY'
    industrial_process_or_use_operation_col = 'Industrial Process or Use Operation'
    cc_product_category_col = 'C/C Product Category'
    industrial_function_category_col = 'Industrial Function Category'
    cc_function_category_col = 'C/C Function Category'
    industrial_other_function_category_col = 'Industrial Other Function Category'
    cc_other_function_category_col = 'C/C Other Function Category'
    cc_other_product_category_col = 'CONS / COMM PROD CAT OTHER DESC'
    recycled_col = 'RECYCLED'
    frs_facility_col = '74. FRS FACILITY ID'
    alternative_frs_facility_col = 'EPA FACILITY REGISTRY ID'
    use_type_col = 'Use Type'
    ind_sites_col = 'INDUSTRIAL SITES'
    ind_sites_code_col = 'IND SITES CODE'
    ind_workers_col = 'WORKERS'
    ind_workers_code_col = 'WORKERS CODE'
    max_conc_col = 'MAXIMUM CONCENTRATION'
    max_conc_code_col = 'MAX CONC CODE'

    naics_cols = [
        '41. NAICS CODE 1', '42. NAICS CODE 2', '43. NAICS CODE 3',
        '44. NAICS CODE 4', '45. NAICS CODE 5', '46. NAICS CODE 6'
    ]

    naics_cols_special = ['NAICS Code 1', 'NAICS Code 2', 'NAICS Code 3']

    matched_naics_cols = [
        'Matched NAICS 1', 'Matched NAICS 2', 'Matched NAICS 3',
        'Matched NAICS 4', 'Matched NAICS 5', 'Matched NAICS 6'
    ]

    matched_naics_desc_cols = [
        'Matched NAICS 1 Description', 'Matched NAICS 2 Description', 'Matched NAICS 3 Description',
        'Matched NAICS 4 Description', 'Matched NAICS 5 Description', 'Matched NAICS 6 Description'
    ]

    matched_product_cols = []
    num_naics_columns = 6

    # Determine if we need to switch to the special case NAICS columns
    all_naics_blank = df[naics_cols].isna().all().all() if all(col in df.columns for col in naics_cols) else True

    if all_naics_blank:
        naics_cols = naics_cols_special
        matched_naics_cols = matched_naics_cols[:3]
        matched_naics_desc_cols = matched_naics_desc_cols[:3]
        num_naics_columns = 3

    # Loop through the possible NAICS columns
    for i in range(1, num_naics_columns + 1):
        column_names = [
            f'Matched NAICS {i} - Product Name',
            f'Matched NAICS {i} - PUC Kind',
            f'Matched NAICS {i} - PUC General Category',
            f'Matched NAICS {i} - PUC Product Family',
            f'Matched NAICS {i} - PUC Product Type',
            f'Matched NAICS {i} - Raw Min Comp',
            f'Matched NAICS {i} - Raw Max Comp',
            f'Matched NAICS {i} - Raw Central Comp',
            f'Matched NAICS {i} - Unit Type',
            f'Matched NAICS {i} - Lower Weight Fraction',
            f'Matched NAICS {i} - Upper Weight Fraction',
            f'Matched NAICS {i} - Central Weight Fraction'
        ]
        matched_product_cols.extend(column_names)

    # Create a dictionary to hold data for each sheet, using sets to avoid duplicates
    sheets_data = {
        "Activity": defaultdict(set),
        "Process or Use Operation": defaultdict(set),
        "C-C Product Use Category": defaultdict(set),
        "Function": defaultdict(set),
        "NAICS": defaultdict(set),
        "Product Use Information": [],
        "End-of-Life": defaultdict(set),
        "Legend": {},
        "Facility Summary": defaultdict(list),
        "TRI Reported Activities": defaultdict(dict)  # New sheet for TRI Reported Activities
    }

    frs_counter = 1
    frs_legend = {}
    frs_code_map = {}

    # Additional data for the legend
    additional_data = {
        'Site Code (S)': ["S0", "S1", "S2", "S3"],
        'Number of Sites': ["Not Known", "< 10", "10 - 24", "25 - 99"],
        'Worker Code (W)': ["W0", "W1", "W2", "W3", "W4", "W5"],
        'Number of Workers': ["Not Known", "< 10", "10 - 24", "25 - 49", "50 - 99", "100 - 499"],
        'Maximum Concentration Code (M)': ["M0", "M1", "M2", "M3", "M4", "M5"],
        'Maximum Concentration (% .wt)': ["Not Known", "< 1%", "1 - < 30%", "30 - < 60%", "60% - <90%", "90% +"],
    }

    # Extracting relevant columns to add into the facility summary
    additional_summary_cols = [
        '2019 DOMESTIC PV', '2019 IMPORT PV', '2019 PV', '2018 PV', '2017 PV', '2016 PV',
        '2019 NATIONALLY AGGREGATED PV', '2018 NATIONALLY AGGREGATED PV',
        '2017 NATIONALLY AGGREGATED PV', '2016 NATIONALLY AGGREGATED PV',
        'IMPORTED CHEM NEVER AT SITE', '2019 V USED ON-SITE', '2019 V EXPORTED'
    ]

    # Iterate over each row in the DataFrame
    for index, row in df.iterrows():
        frs_id = row[frs_facility_col] if pd.notna(row[frs_facility_col]) else row[alternative_frs_facility_col]
        if pd.isna(frs_id):
            continue

        # Ensure Use Type is properly retrieved
        use_type = row[use_type_col] if use_type_col in row and pd.notna(row[use_type_col]) else ''
        use_type_suffix = '(C)' if use_type == 'Consumer and Commercial' else '(I)' if use_type == 'Industrial' else ''

        if frs_id not in frs_code_map:
            frs_code = f'FRS{frs_counter}{use_type_suffix}'
            frs_code_map[frs_id] = frs_code
            frs_legend[frs_code] = frs_id
            frs_counter += 1
        else:
            frs_code = frs_code_map[frs_id]
        
        # Update FRS code suffix if Use Type is determined later
        if frs_code.endswith('(C)') or frs_code.endswith('(I)'):
            if use_type_suffix and not frs_code.endswith(use_type_suffix):
                frs_code = frs_code[:-3] + use_type_suffix  # Replace existing suffix with the correct one
                frs_code_map[frs_id] = frs_code
                frs_legend[frs_code] = frs_id


        # Manufacturing/Import
        activity = row[activity_col] if pd.notna(row[activity_col]) else 'CBI'
        if activity in ['Import', 'CBI', 'Manufacture']:
            sheets_data['Activity'][activity].add(frs_code)

        # Process or Use Operation
        industrial_process_or_use_operation = row[industrial_process_or_use_operation_col] if industrial_process_or_use_operation_col in row and pd.notna(row[industrial_process_or_use_operation_col]) else None
        # Consumer and Commercial Product Category
        cc_product_category = row[cc_product_category_col] if cc_product_category_col in row and pd.notna(row[cc_product_category_col]) else None

        # Handle 'Other' and 'Other (specify)' cases for Industrial Process or Use Operation
        if industrial_process_or_use_operation in ['Other (specify)', 'Other']:
            industrial_other_process_or_use_operation_col = 'INDUSTRIAL OTHER PROCESS OR USE OPERATION'
            if industrial_other_process_or_use_operation_col in row:
                industrial_process_or_use_operation = row[industrial_other_process_or_use_operation_col]
        if pd.notna(industrial_process_or_use_operation):
            sheets_data['Process or Use Operation'][industrial_process_or_use_operation].add(frs_code)

        #Product Use Category
        # Handle 'Other' and 'Other (specify)' cases for C/C Product Category
        if cc_product_category in ['Other (specify)', 'Other']:
            cc_other_product_category_col = 'CONS / COMM PROD CAT OTHER DESC'
            if cc_other_product_category_col in row:
                cc_product_category = row[cc_other_product_category_col]
        if pd.notna(cc_product_category):
            sheets_data['C-C Product Use Category'][cc_product_category].add(frs_code)


        # Function - Handle Industrial and C/C function categories
        industrial_function_category = row[industrial_function_category_col] if pd.notna(row[industrial_function_category_col]) else None
        cc_function_category = row[cc_function_category_col] if pd.notna(row[cc_function_category_col]) else None

        # Handle 'Other' and 'Other (specify)' cases for Industrial
        if industrial_function_category in ['Other (specify)', 'Other']:
            industrial_other_function_category_col = 'Industrial Other Function Category'
            if industrial_other_function_category_col in row:
                industrial_function_category = row[industrial_other_function_category_col]
        if pd.notna(industrial_function_category):
            sheets_data['Function'][industrial_function_category].add(frs_code)

        # Handle 'Other' and 'Other (specify)' cases for C/C
        if cc_function_category in ['Other (specify)', 'Other']:
            cc_other_function_category_col = 'C/C Other Function Category'
            if cc_other_function_category_col in row:
                cc_function_category = row[cc_other_function_category_col]
        if pd.notna(cc_function_category):
            sheets_data['Function'][cc_function_category].add(frs_code)



        # NAICS
        for naics_col, matched_naics_col, matched_naics_desc_col in zip(naics_cols, matched_naics_cols, matched_naics_desc_cols):
            if naics_col in row:
                naics_code = row[naics_col]
                matched_naics = row.get(matched_naics_col)
                matched_naics_desc = row.get(matched_naics_desc_col)
                if pd.notna(naics_code):
                    naics_code_str = str(naics_code)[:6]  # Truncate to first 6 characters
                    sheets_data['NAICS'][naics_code_str].add(frs_code)
                if pd.notna(matched_naics) and pd.notna(matched_naics_desc):
                    matched_naics_str = str(matched_naics)
                    matched_naics_desc_str = str(matched_naics_desc)

                    sheets_data['NAICS'][f'{matched_naics_str} - {matched_naics_desc_str}'].add(frs_code)

        # Product Use Information
        for i in range(1, num_naics_columns + 1):
            naics_col = f'Matched NAICS {i}'
            naics_desc_col = f'Matched NAICS {i} Description'
            product_name_col = f'Matched NAICS {i} - Product Name'
            if pd.notna(row.get(product_name_col)):
                product_details = [
                    row[product_name_col],
                    frs_code,
                    row.get(naics_col),
                    row.get(naics_desc_col),
                    row.get(f'Matched NAICS {i} - PUC Kind', ''),
                    row.get(f'Matched NAICS {i} - PUC General Category', ''),
                    row.get(f'Matched NAICS {i} - PUC Product Family', ''),
                    row.get(f'Matched NAICS {i} - PUC Product Type', ''),
                    row.get(f'Matched NAICS {i} - Raw Min Comp', ''),
                    row.get(f'Matched NAICS {i} - Raw Max Comp', ''),
                    row.get(f'Matched NAICS {i} - Raw Central Comp', ''),
                    row.get(f'Matched NAICS {i} - Unit Type', ''),
                    row.get(f'Matched NAICS {i} - Lower Weight Fraction', ''),
                    row.get(f'Matched NAICS {i} - Upper Weight Fraction', ''),
                    row.get(f'Matched NAICS {i} - Central Weight Fraction', '')
                ]
                sheets_data['Product Use Information'].append(product_details)

        # End-of-Life
        recycled = row[recycled_col]
        if recycled in ['Yes', 'No', 'CBI']:
            sheets_data['End-of-Life'][recycled].add(frs_code)

        # Facility Summary
        ind_sites = row.get(ind_sites_col, 'NKRA')
        ind_sites_code = row.get(ind_sites_code_col, 'S0')
        if pd.isna(ind_sites) or ind_sites == 'Not Known or Reasonably Ascertainable':
            ind_sites = 'NKRA'
            ind_sites_code = 'S0'

        ind_workers = row.get(ind_workers_col, 'NKRA')
        ind_workers_code = row.get(ind_workers_code_col, 'W0')
        if pd.isna(ind_workers) or ind_workers == 'NKRA':
            ind_workers = 'NKRA'
            ind_workers_code = 'W0'

        max_conc = row.get(max_conc_col, 'NKRA')
        max_conc_code = row.get(max_conc_code_col, 'M0')
        if pd.isna(max_conc) or max_conc == 'NKRA':
            max_conc = 'NKRA'
            max_conc_code = 'M0'

        # Extract additional columns
        additional_summary_values = [row.get(col, '') for col in additional_summary_cols]

        sheets_data['Facility Summary'][frs_code].append([
            frs_id,
            ind_sites,
            ind_sites_code,
            ind_workers,
            ind_workers_code,
            max_conc,
            max_conc_code
        ] + additional_summary_values)

    # Convert sets back to lists for DataFrame creation
    for sheet in sheets_data:
        if isinstance(sheets_data[sheet], defaultdict):
            for key in sheets_data[sheet]:
                sheets_data[sheet][key] = list(sheets_data[sheet][key])

    # Process TRI reported activities
    tri_reported_activities_cols = [
        '89. PRODUCE THE CHEMICAL', '90. IMPORT THE CHEMICAL', '91. ON-SITE USE OF THE CHEMICAL',
        '92. SALE OR DISTRIBUTION OF THE CHEMICAL', '93. AS A BYPRODUCT', '94. AS A MANUFACTURED IMPURITY',
        '95. USED AS A REACTANT', '96. P101  FEEDSTOCKS', '97. P102  RAW MATERIALS', '98. P103  INTERMEDIATES',
        '99. P104  INITIATORS', '100. P199  OTHER', '101. ADDED AS A FORMULATION COMPONENT', '102. P201  ADDITIVES',
        '103. P202  DYES', '104. P203  REACTION DILUENTS', '105. P204  INITIATORS', '106. P205  SOLVENTS',
        '107. P206  INHIBITORS', '108. P207  EMULSIFIERS', '109. P208  SURFACTANTS', '110. P209  LUBRICANTS',
        '111. P210  FLAME RETARDANTS', '112. P211  RHEOLOGICAL MODIFIERS', '113. P299  OTHER',
        '114. USED AS AN ARTICLE COMPONENT', '115. REPACKAGING', '116. AS A PROCESS IMPURITY', '117. PROCESSED / RECYCLING',
        '118. USED AS A CHEMICAL PROCESSING AID', '119. Z101  PROCESS SOLVENTS', '120. Z102  CATALYSTS',
        '121. Z103  INHIBITORS', '122. Z104  INITIATORS', '123. Z105  REACTION TERMINATORS', '124. Z106  SOLUTION BUFFERS',
        '125. Z199  OTHER', '126. USED AS A MANUFACTURING AID', '127. Z201  PROCESS LUBRICANTS', '128. Z202  METALWORKING FLUIDS',
        '129. Z203  COOLANTS', '130. Z204  REFRIGERANTS', '131. Z205  HYDRAULIC FLUIDS', '132. Z299  OTHER',
        '133. ANCILLARY OR OTHER USE', '134. Z301  CLEANER', '135. Z302  DEGREASER', '136. Z303  LUBRICANT',
        '137. Z304  FUEL', '138. Z305  FLAME RETARDANT', '139. Z306  WASTE TREATMENT', '140. Z307  WATER TREATMENT',
        '141. Z308  CONSTRUCTION MATERIALS', '142. Z399  OTHER'
    ]

    tri_reported_activities_data = {}

    for index, row in df.iterrows():
        frs_id = row[frs_facility_col] if pd.notna(row[frs_facility_col]) else row[alternative_frs_facility_col]
        if pd.isna(frs_id):
            continue

        if frs_id in frs_code_map:
            frs_code = frs_code_map[frs_id]
            if frs_code not in tri_reported_activities_data:
                tri_reported_activities_data[frs_code] = {
                    'FRS Code': frs_code,
                    'Actual FRS ID': frs_id
                }

            for col in tri_reported_activities_cols:
                tri_reported_activities_data[frs_code][col] = row[col]

    # Convert dictionary to DataFrame
    tri_reported_activities_df = pd.DataFrame(list(tri_reported_activities_data.values()))

    # Store TRI Reported Activities data in sheets_data
    sheets_data['TRI Reported Activities'] = tri_reported_activities_data

    # Create a DataFrame for each sheet and write to Excel
    output_excel_path = f'qualitative_chemical_flow_mapping_summary_{cas_number}.xlsx'
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        for sheet_name, data in sheets_data.items():
            if sheet_name == 'Legend':
                legend_df = pd.DataFrame(list(frs_legend.items()), columns=['FRS Code', 'Actual FRS ID'])
                # Adding additional legend data to the DataFrame
                for key, values in additional_data.items():
                    legend_df[key] = pd.Series(values)
                legend_df.to_excel(writer, sheet_name=sheet_name, index=False)

            elif sheet_name == 'Facility Summary':
                summary_data = []
                for frs_code, details in data.items():
                    summary_data.append([
                        frs_code,
                        details[0][0],  # Actual FRS ID
                        details[0][1],  # Number of Sites (S)
                        details[0][2],  # Site Code
                        details[0][3],  # Number of Workers (W)
                        details[0][4],  # Worker Code
                        details[0][5],  # Maximum Concentration (M in % wt.)
                        details[0][6]   # Maximum Concentration Code
                    ] + details[0][7:])  # Adding the additional columns here
                summary_df = pd.DataFrame(summary_data, columns=[
                    'FRS Code', 'Actual FRS ID', 'Number of Sites (S)', 'Site Code', 'Number of Workers (W)', 'Worker Code',
                    'Maximum Concentration (M in % wt.)', 'Maximum Concentration Code'] + additional_summary_cols)
                
                summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
            elif sheet_name == 'TRI Reported Activities':
                tri_reported_activities_df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Access the worksheet
                sheet = writer.sheets[sheet_name]

                # Define the fill and font styles
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                dark_green_font = Font(color='006100')

                # Apply conditional formatting
                for row_idx in range(2, sheet.max_row + 1):  # Start from the second row (excluding header)
                    for col_idx in range(2, sheet.max_column + 1):  # Start from the second column (excluding FRS Code and Actual FRS ID)
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if cell.value == 'Yes':
                            cell.fill = green_fill
                            cell.font = dark_green_font

            elif sheet_name == 'Product Use Information':
                product_use_category_df = pd.DataFrame(data, columns=[
                    'Product Name', 'FRS Code', 'Matched NAICS', 'Matched NAICS Description',
                    'PUC Kind', 'PUC General Category', 'PUC Product Family', 'PUC Product Type',
                    'Raw Min Comp', 'Raw Max Comp', 'Raw Central Comp', 'Unit Type',
                    'Lower Weight Fraction', 'Upper Weight Fraction', 'Central Weight Fraction'
                ])
                product_use_category_df.to_excel(writer, sheet_name='Product Use Information', index=False)  # Renamed to Product Use Information

            else:
                sheet_df = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in data.items()]))
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Add the "NAICS Description" sheet
        naics_df.to_excel(writer, sheet_name='NAICS Description', index=False)

        # Truncate NAICS codes to the first 6 characters in the header row of the 'NAICS' sheet and convert to string
        naics_sheet = writer.sheets['NAICS']
        for col in range(2, naics_sheet.max_column + 1):
            cell = naics_sheet.cell(row=1, column=col)  # Header row is row 1
            if cell.value:
                cell.value = str(cell.value)[:6]  # Convert to string and truncate to first 6 characters

    print(f"Excel file '{output_excel_path}' created successfully.")
    return output_excel_path




# Step 5: Generate report and diagram
def generate_report_and_diagram(qualitative_summary_file, casrn):
    df_facility_summary = pd.read_excel(qualitative_summary_file, sheet_name='Facility Summary')
    df_tri_reported_activities = pd.read_excel(qualitative_summary_file, sheet_name='TRI Reported Activities')
    df_legend = pd.read_excel(qualitative_summary_file, sheet_name='Legend').iloc[:6]
    df_naics_description = pd.read_excel(qualitative_summary_file, sheet_name='NAICS Description')

    # Filtered FRS codes based on entries in the Process or Use Operation tab and C-C Product Use Category tab
    df_process = pd.read_excel(qualitative_summary_file, sheet_name='Process or Use Operation')
    df_cc_product_use_category = pd.read_excel(qualitative_summary_file, sheet_name='C-C Product Use Category')

    # Combine the FRS codes from both sheets
    valid_frs_codes_process = set(df_process.values.flatten()) - {None}
    valid_frs_codes_cc_product_use = set(df_cc_product_use_category.values.flatten()) - {None}

    # Union of both sets to get all valid FRS codes
    valid_frs_codes = valid_frs_codes_process.union(valid_frs_codes_cc_product_use)


    # Define the combine_naics function
    def combine_naics(df):
        combined = {}
        for col in df.columns:
            key = col[:6]
            if key not in combined:
                combined[key] = []
            combined[key].extend(df[col].dropna().tolist())
        combined_df = pd.DataFrame.from_dict(combined, orient='index').transpose()
        return combined_df

    # Add NAICS descriptions to labels
    def add_naics_descriptions(df, df_naics_description):
        descriptions = {}
        for naics_code in df.columns:
            if naics_code == "CBI":
                descriptions[naics_code] = "Confidential Business Information"
            else:
                try:
                    description = df_naics_description[df_naics_description['2022 NAICS US   Code'] == int(naics_code)]
                    if not description.empty:
                        descriptions[naics_code] = description['2022 NAICS US Title'].values[0]
                    else:
                        descriptions[naics_code] = ""
                except ValueError:
                    descriptions[naics_code] = ""
        return descriptions


    # Combine NAICS data
    df_naics = pd.read_excel(qualitative_summary_file, sheet_name='NAICS')
    df_naics_combined = combine_naics(df_naics)

    # Get NAICS descriptions
    naics_descriptions = add_naics_descriptions(df_naics_combined, df_naics_description)

    # Select the required columns
    columns = ['FRS Code', 'Actual FRS ID', 'Number of Sites (S)', 'Site Code', 'Number of Workers (W)', 'Worker Code']
    max_concentration_columns = ['FRS Code', 'Actual FRS ID', 'Maximum Concentration (M in % wt.)', 'Maximum Concentration Code']
    pv_columns = ['FRS Code', 'Actual FRS ID', '2019 DOMESTIC PV', '2019 IMPORT PV', '2019 PV', '2018 PV', '2017 PV', '2016 PV']
    nationally_aggregated_columns_part1 = ['FRS Code', 'Actual FRS ID', '2019 NATIONALLY AGGREGATED PV', '2018 NATIONALLY AGGREGATED PV']
    nationally_aggregated_columns_part2 = ['FRS Code', 'Actual FRS ID', '2017 NATIONALLY AGGREGATED PV', '2016 NATIONALLY AGGREGATED PV']
    chemical_presence = ['FRS Code', 'Actual FRS ID', 'IMPORTED CHEM NEVER AT SITE', '2019 V USED ON-SITE', '2019 V EXPORTED']

    # Filter the dataframe based on the valid FRS codes
    df_filtered = df_facility_summary[df_facility_summary['FRS Code'].isin(valid_frs_codes)][columns]
    df_max_concentration = df_facility_summary[df_facility_summary['FRS Code'].isin(valid_frs_codes)][max_concentration_columns]
    df_pv = df_facility_summary[df_facility_summary['FRS Code'].isin(valid_frs_codes)][pv_columns]
    df_nationally_aggregated_p1 = df_facility_summary[df_facility_summary['FRS Code'].isin(valid_frs_codes)][nationally_aggregated_columns_part1]
    df_nationally_aggregated_p2 = df_facility_summary[df_facility_summary['FRS Code'].isin(valid_frs_codes)][nationally_aggregated_columns_part2]
    df_chemical_presence = df_facility_summary[df_facility_summary['FRS Code'].isin(valid_frs_codes)][chemical_presence]

    # Manually create dataframes for TRI Reported Activities with 3 columns per part
    tri_parts = []
    for i in range(0, len(df_tri_reported_activities.columns[2:]), 3):
        tri_part_columns = ['FRS Code', 'Actual FRS ID'] + list(df_tri_reported_activities.columns[2+i:5+i])
        tri_parts.append(df_tri_reported_activities[df_tri_reported_activities['FRS Code'].isin(valid_frs_codes)][tri_part_columns])

    # Convert 'Actual FRS ID' to non-scientific notation using .loc to avoid SettingWithCopyWarning
    df_filtered.loc[:, 'Actual FRS ID'] = df_filtered['Actual FRS ID'].apply(lambda x: f"{float(x):.0f}" if isinstance(x, (int, float)) else x)
    df_max_concentration.loc[:, 'Actual FRS ID'] = df_max_concentration['Actual FRS ID'].apply(lambda x: f"{float(x):.0f}" if isinstance(x, (int, float)) else x)
    df_pv.loc[:, 'Actual FRS ID'] = df_pv['Actual FRS ID'].apply(lambda x: f"{float(x):.0f}" if isinstance(x, (int, float)) else x)
    df_nationally_aggregated_p1.loc[:, 'Actual FRS ID'] = df_nationally_aggregated_p1['Actual FRS ID'].apply(lambda x: f"{float(x):.0f}" if isinstance(x, (int, float)) else x)
    df_nationally_aggregated_p2.loc[:, 'Actual FRS ID'] = df_nationally_aggregated_p2['Actual FRS ID'].apply(lambda x: f"{float(x):.0f}" if isinstance(x, (int, float)) else x)
    df_chemical_presence.loc[:, 'Actual FRS ID'] = df_chemical_presence['Actual FRS ID'].apply(lambda x: f"{float(x):.0f}" if isinstance(x, (int, float)) else x)
    for tri_part in tri_parts:
        tri_part.loc[:, 'Actual FRS ID'] = tri_part['Actual FRS ID'].apply(lambda x: f"{float(x):.0f}" if isinstance(x, (int, float)) else x)

    # Replace NaN values with "Unknown" in df_chemical_presence and "No" in TRI parts
    df_chemical_presence = df_chemical_presence.fillna("Unknown")
    for i in range(len(tri_parts)):
        tri_parts[i] = tri_parts[i].fillna("No")


    # Function to split header into a specified number of lines while preserving word order
    def split_header(header, max_lines=3):
        words = header.split()
        if len(words) <= max_lines:
            return '\n'.join(words)
        # Distribute words evenly across lines while maintaining order
        n = len(words)
        k = (n + max_lines - 1) // max_lines
        return '\n'.join(' '.join(words[i:i+k]) for i in range(0, n, k))

    # PDF setup
    page_width, page_height = portrait(letter)
    margin = 0.8 * inch  # Reduced margin to push content up
    usable_width = page_width - 2 * margin
    usable_height = page_height - 2 * margin
    max_rows_per_page = 26  # Maximum number of rows per page

    # Create the PDF
    pdf_path = f'facility_summary_report_{casrn}.pdf'
    c = canvas.Canvas(pdf_path, pagesize=portrait(letter))

    # Title
    main_title = f"Facility Summary Report (CAS No. {casrn})"
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(page_width / 2, page_height - margin / 2, main_title)

    # Function to draw a table with multiline headers (generating reports)
    def draw_table(c, df, columns, x, y, width, height, title, max_header_lines=3, max_rows=max_rows_per_page):
        def draw_page_title():
            c.setFont("Helvetica-Bold", 16)
            c.drawCentredString(page_width / 2, page_height - margin / 2, main_title)
            c.setFont("Helvetica-Bold", 14)
            c.drawString(x, y + 20, title)  # Adjusted title position
        
        draw_page_title()  # Draw the title on the first page

        # Split column headers into multiple lines if necessary
        header_lines = [split_header(col, max_lines=max_header_lines) for col in columns]
        
        # Prepare data and column widths
        data = [header_lines] + df[columns].values.tolist()
        col_widths = [max(df[col].astype(str).apply(len).max(), len(col)) * 5.3 for col in columns]  # Estimate column widths

        # Split data into chunks for pagination
        num_pages = (len(data) - 1 + max_rows) // max_rows  # Calculate the number of pages needed
        for page in range(num_pages):
            start_row = page * max_rows
            end_row = start_row + max_rows
            page_data = [header_lines] + data[start_row + 1:end_row + 1]
            table = Table(page_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ]))
            table.wrapOn(c, width, height)
            table.drawOn(c, x, y - 30 - table._height)  # Adjusted table position

            if page < num_pages - 1:
                c.showPage()
                draw_page_title()  # Draw the title on each continuation page

    # Draw the tables with pagination
    draw_table(c, df_filtered, columns, margin, page_height - margin - 60, usable_width, usable_height, "Facility Summary")  # Adjusted position
    c.showPage()
    draw_table(c, df_max_concentration, max_concentration_columns, margin, page_height - margin - 60, usable_width, usable_height, "Facility Summary")
    c.showPage()
    draw_table(c, df_pv, pv_columns, margin, page_height - margin - 60, usable_width, usable_height, "Facility Summary")
    c.showPage()
    draw_table(c, df_nationally_aggregated_p1, nationally_aggregated_columns_part1, margin, page_height - margin - 60, usable_width, usable_height, "Facility Summary")
    c.showPage()
    draw_table(c, df_nationally_aggregated_p2, nationally_aggregated_columns_part2, margin, page_height - margin - 60, usable_width, usable_height, "Facility Summary")
    c.showPage()
    draw_table(c, df_chemical_presence, chemical_presence, margin, page_height - margin - 60, usable_width, usable_height, "Facility Summary")
    c.showPage()

    # Draw the TRI Reported Activities tables on new pages
    for tri_part in tri_parts:
        # Remove numbers from column names
        tri_part.columns = [re.sub(r'^\d+\.\s*', '', col) for col in tri_part.columns]
        draw_table(c, tri_part, tri_part.columns, margin, page_height - margin - 60, usable_width, usable_height, "TRI Reported Activities")
        c.showPage()

    # Draw the first Legend table on a new page
    legend_columns_part1 = ['Site Code (S)', 'Number of Sites', 'Worker Code (W)', 'Number of Workers']
    df_legend_part1 = df_legend[legend_columns_part1]
    df_legend_part1 = df_legend_part1.fillna("")

    draw_table(c, df_legend_part1, legend_columns_part1, margin, page_height - margin - 60, usable_width, usable_height, "Legend")
    c.showPage()

    # Draw the second Legend table on a new page
    legend_columns_part2 = ['Maximum Concentration Code (M)', 'Maximum Concentration (% .wt)']
    df_legend_part2 = df_legend[legend_columns_part2]
    df_legend_part2 = df_legend_part2.fillna("")

    draw_table(c, df_legend_part2, legend_columns_part2, margin, page_height - margin - 60, usable_width, usable_height, "Legend")

    # Add Use Type legend text below the second table
    c.setFont("Helvetica", 12)
    legend_text_y_position = page_height - margin - 60 - df_legend_part2.shape[0] * 15 - 120  # Adjust the value as needed
    c.drawString(margin, legend_text_y_position, "Use Type Codes: (I) Industrial / (C) Consumer and Commercial")

    c.showPage()
    c.save()

    print(f"PDF file '{pdf_path}' created successfully.")

    # Flow Diagram Generation
    df_activity = pd.read_excel(qualitative_summary_file, sheet_name='Activity')
    df_function = pd.read_excel(qualitative_summary_file, sheet_name='Function')
    df_naics = pd.read_excel(qualitative_summary_file, sheet_name='NAICS')
    df_end_of_life = pd.read_excel(qualitative_summary_file, sheet_name='End-of-Life')
    df_product_use = pd.read_excel(qualitative_summary_file, sheet_name='Product Use Information')

    # Initialize a set to track unique connections
    unique_connections = set()

    def add_unique_connection(source_node, target_node, edge_style, unique_connections):
        connection = (source_node, target_node)
        if connection not in unique_connections:
            source_node >> edge_style >> target_node
            unique_connections.add(connection)

    # Create custom nodes for FRS codes
    def create_frs_node(label):
        split_label_text = split_label(label, max_length=20)
        return Custom(split_label_text, "./frs_icon.png", fontsize="25")  # Adjust fontsize for FRS labels

    # Define the diagram with spacing and font size adjustments
    graph_attr = {
        "splines": "polyline",  # Use straight lines
        "nodesep": "0.2",  # Adjust node separation
        "ranksep": "6",  # Adjust rank separation
        "fontsize": "60",  # Font size for the diagram
        "fontname": "Arial-BoldMT",  # Bold font for the diagram
        "labeljust": "c",  # Center the label
        "labelloc": "t",  # Place the label at the top
        "label": f"Qualitative Mapping (CAS No. {casrn})",  # Title with CAS number
        "fontcolor": "black",
        "fontsize": "80",  # Font size of the title
        "rankdir": "TB"  # sets the direction from top to bottom
    }

    node_attr = {
        "fontsize": "48",  # font size for FRS nodes
        "fontname": "Arial-BoldMT"  # Font for FRS nodes
    }

    cluster_attr = {
        "fontsize": "84",  # font size for cluster labels
        "fontname": "Arial-BoldMT",  # Bold font for cluster labels
        "fontcolor": "black",  # Font color
        "width": "4",  # Width of the cluster
        "height": "2",  # Height of the cluster
        "labeljust": "c",  # Center the label horizontally
        "labelloc": "c"  # Center the label vertically
    }

    section_attr = {
        "fontsize": "64",  # font size for section labels
        "fontname": "Arial-BoldMT",  # Bold font for section labels
        "fontcolor": "black",  # Font color
        "width": "4",  # Width of the section
        "height": "2",  # Height of the section
        "labeljust": "c",  # Center the label horizontally
        "labelloc": "c"  # Center the label vertically
    }

    # Create custom nodes for FRS codes with split labels
    def create_frs_node(label):
        split_label_text = split_label(label, max_length=20)
        return Custom(split_label_text, "./frs_icon.png", fontsize="25")  # Adjust fontsize for FRS labels

    def create_frs_nodes_vertically(frs_list):
        nodes = []
        for frs in frs_list:
            if frs in valid_frs_codes:
                nodes.append(create_frs_node(frs))
        return nodes

    def create_frs_nodes_multiline(frs_list, max_per_line=3):
        nodes = []
        line_nodes = []
        count = 0
        
        for frs in frs_list:
            if frs in valid_frs_codes:
                node = create_frs_node(frs)
                line_nodes.append(node)
                count += 1
                
                if count == max_per_line:
                    nodes.extend(line_nodes)
                    line_nodes = []
                    count = 0
        
        if line_nodes:
            nodes.extend(line_nodes)
        
        return nodes

    # Combine NAICS sections based on the first 6 digits
    def combine_naics(df):
        combined = {}
        for col in df.columns:
            key = col[:6]
            if key not in combined:
                combined[key] = []
            combined[key].extend(df[col].dropna().tolist())
        combined_df = pd.DataFrame.from_dict(combined, orient='index').transpose()
        return combined_df

    df_naics_combined = combine_naics(df_naics)

    # Extract unique product names
    unique_product_names = df_product_use['Product Name'].unique()


    # Function to split label into multiple lines if it exceeds a certain length
    def split_label(label, max_length=20):
        words = label.split()
        lines = []
        current_line = ""
        for word in words:
            if len(current_line) + len(word) + 1 <= max_length:
                current_line += " " + word
            else:
                lines.append(current_line.strip())
                current_line = word
        lines.append(current_line.strip())
        return "\n".join(lines)

    # Create and save the flow diagram using Diagram module
    diagram_path = f'qualitative_mapping_{casrn}'

    with Diagram("", show=False, direction="TB", filename=diagram_path, outformat="png", graph_attr=graph_attr):
        # Manufacturing/Import Cluster
        cbi = []
        manufacture = []
        import_ = []

        with Cluster("Manufacturing/Import", graph_attr={**cluster_attr, "style": "filled", "bgcolor": "lightblue"}) as manufacturing_import_cluster:
            with Cluster("Confidential Business Information", graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                if 'CBI' in df_activity.columns:
                    cbi_frs = df_activity['CBI'].dropna()
                    cbi.extend(create_frs_nodes_multiline(cbi_frs))
            with Cluster("Manufacture", graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                if 'Manufacture' in df_activity.columns:
                    manufacture_frs = df_activity['Manufacture'].dropna()
                    manufacture.extend(create_frs_nodes_multiline(manufacture_frs))
            with Cluster("Import", graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                if 'Import' in df_activity.columns:
                    import_frs = df_activity['Import'].dropna()
                    import_.extend(create_frs_nodes_multiline(import_frs))

        # Process or Use Operation Cluster
        process_sections = {}
        with Cluster("Industrial Process or Use Operation", graph_attr={**cluster_attr, "style": "filled", "bgcolor": "lightgreen"}):
            for col in df_process.columns:
                col_label = split_label(col, max_length=30)
                with Cluster(col_label, graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                    process_frs = df_process[col].dropna()
                    process_sections[col] = create_frs_nodes_multiline(process_frs)

        # C-C Product Use Category Cluster
        cc_product_use_sections = {}
        with Cluster("Consumer and Commercial Product Use Category", graph_attr={**cluster_attr, "style": "filled", "bgcolor": "lightgreen"}) as cc_product_use_category_cluster:
            for col in df_cc_product_use_category.columns:
                col_label = split_label(col, max_length=30)
                with Cluster(col_label, graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                    cc_product_use_frs = df_cc_product_use_category[col].dropna()
                    cc_product_use_sections[col] = create_frs_nodes_multiline(cc_product_use_frs)
            

        # Function Cluster
        function_sections = {}
        with Cluster("Function", graph_attr={**cluster_attr, "style": "filled", "bgcolor": "lightcyan"}):
            for col in df_function.columns:
                col_label = split_label(col, max_length=30)
                with Cluster(col_label, graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                    function_frs = df_function[col].dropna()
                    function_sections[col] = create_frs_nodes_multiline(function_frs)

        # NAICS Cluster
        naics_sections = {}
        with Cluster("NAICS Codes and Descriptions", graph_attr={**cluster_attr, "style": "filled", "bgcolor": "lightgoldenrodyellow"}):
            for col in df_naics_combined.columns:
                naics_label = split_label(col + "\n" + naics_descriptions.get(col, ""), max_length=30)
                with Cluster(naics_label, graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                    naics_frs = df_naics_combined[col].dropna()
                    naics_sections[col] = create_frs_nodes_multiline(naics_frs)

        # Product Use Information Cluster
        product_use_sections = {}
        with Cluster("Product Use Information", graph_attr={**cluster_attr, "style": "filled", "bgcolor": "lightgrey"}):
            if unique_product_names.size > 0:
                for product in unique_product_names:
                    product_label = split_label(product, max_length=25)
                    with Cluster(product_label, graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                        product_use_sections[product] = {}
                        product_df = df_product_use[df_product_use['Product Name'] == product]
                        for naics in product_df['Matched NAICS'].unique():
                            naics_label = split_label(f"NAICS {naics}", max_length=20)
                            with Cluster(naics_label, graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                                naics_frs = product_df[product_df['Matched NAICS'] == naics]['FRS Code'].dropna()
                                product_use_sections[product][naics] = create_frs_nodes_multiline(naics_frs)
            else:
                no_product_info_node = Custom("No Product Information Found for this Chemical", "", shape="box", fontsize="25")
                product_use_sections['No Product Information'] = [no_product_info_node]

        # End-of-Life Cluster
        end_of_life_sections = {}
        with Cluster("Industrial End-of-Life", graph_attr={**cluster_attr, "style": "filled", "bgcolor": "lightpink"}):
            with Cluster("Confidential Business Information", graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                if 'CBI' in df_end_of_life.columns:
                    eol_cbi_frs = df_end_of_life['CBI'].dropna()
                    end_of_life_sections['CBI'] = create_frs_nodes_multiline(eol_cbi_frs)
            with Cluster("Recycling", graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                if 'Yes' in df_end_of_life.columns:
                    eol_yes_frs = df_end_of_life['Yes'].dropna()
                    end_of_life_sections['Recycling'] = create_frs_nodes_multiline(eol_yes_frs)
            with Cluster("Disposal", graph_attr={**section_attr, "style": "filled", "bgcolor": "white"}):
                if 'No' in df_end_of_life.columns:
                    eol_no_frs = df_end_of_life['No'].dropna()
                    end_of_life_sections['Disposal'] = create_frs_nodes_multiline(eol_no_frs)

        # Creating representative nodes for cluster connections with different colors for each transition
        mi_to_process = Custom(" ", "", shape="circle", style="filled", color="lightgreen")
        mi_to_cc_product_use = Custom(" ", "", shape="circle", style="filled", color="lightseagreen")
        convergence_to_function = Custom(" ", "", shape="circle", style="filled", color="lightcyan")
        function_to_eol = Custom(" ", "", shape="circle", style="filled", color="lightpink")
        function_to_naics = Custom(" ", "", shape="circle", style="filled", color="lightgoldenrodyellow")
        naics_to_product_use = Custom(" ", "", shape="circle", style="filled", color="lightgray")
        eol_to_mi = Custom(" ", "", shape="circle", style="filled", color="lightblue") if 'Recycling' in end_of_life_sections and end_of_life_sections['Recycling'] else None

        # Custom edge style
        custom_edge = Edge(color="black", style="bold", penwidth="2.0")
        custom_edge_product_use_to_function = Edge(color="darkblue", style="bold", penwidth="2.0")
        custom_edge_recycled = Edge(color="darkgreen", style="bold", penwidth="2.0")
        custom_edge_eol = Edge(color="pink", style ="bold", penwidth="2.0")

        # Connecting Clusters with custom arrows
        for node in cbi + manufacture + import_:
            node >> custom_edge >> mi_to_process
            node >> custom_edge >> mi_to_cc_product_use

        # Connect Manufacturing/Import sections to Process sections
        for section, nodes in process_sections.items():
            for node in nodes:
                add_unique_connection(mi_to_process, node, custom_edge, unique_connections)

        # Connect Manufacturing/Import sections to C-C Product Use Category sections
        for section, nodes in cc_product_use_sections.items():
            for node in nodes:
                add_unique_connection(mi_to_cc_product_use, node, custom_edge, unique_connections)

        # Connect Process sections to the convergence node
        for section, nodes in process_sections.items():
            for node in nodes:
                node >> custom_edge >> convergence_to_function

        # Connect C-C Product Use Category sections to the convergence node
        for section, nodes in cc_product_use_sections.items():
            for node in nodes:
                node >> custom_edge >> convergence_to_function

        # Connect the convergence node to Function sections
        for section, nodes in function_sections.items():
            for node in nodes:
                convergence_to_function >> custom_edge >> node


        for section, nodes in function_sections.items():
            for node in nodes:
                node >> custom_edge_eol >> function_to_eol
                node >> custom_edge >> function_to_naics

        for section, nodes in end_of_life_sections.items():
            if nodes:
                function_to_eol >> custom_edge_eol >> nodes[0]

        for section, nodes in naics_sections.items():
            if nodes:
                function_to_naics >> custom_edge >> nodes[0]

        for naics_section, naics_nodes in naics_sections.items():
            for node in naics_nodes:
                node >> custom_edge >> naics_to_product_use

        # Connecting NAICS sections to Product Use Information
        for product, naics_dict in product_use_sections.items():
            if isinstance(naics_dict, dict):  # Check if naics_dict is a dictionary
                for naics, frs_nodes in naics_dict.items():
                    for node in frs_nodes:
                        naics_to_product_use >> node
            else:
                for node in naics_dict:  # Handle the case where naics_dict is not a dictionary
                    naics_to_product_use >> node

        # Adding arrows from Recycling to Manufacturing/Import
        if 'Recycling' in end_of_life_sections and end_of_life_sections['Recycling']:
            for node in end_of_life_sections['Recycling']:
                node >> custom_edge_recycled >> eol_to_mi

            if eol_to_mi:  # Ensure eol_to_mi node exists before connecting
                if cbi:
                    eol_to_mi >> custom_edge_recycled >> cbi[0]
                elif manufacture:
                    eol_to_mi >> custom_edge_recycled >> manufacture[0]
                elif import_:
                    eol_to_mi >> custom_edge_recycled >> import_[0]




    print(f"Diagram file 'qualitative_mapping_{casrn}.png' created successfully.")


# Helper function to filter FRS codes and print debug statements
def filter_frs_codes(df, valid_frs_codes, sheet_name):

    filtered_df = df.applymap(lambda x: x if x in valid_frs_codes else None)
    
    return filtered_df



#Step 6 - Simplified Diagram Generation.
def generate_simplified_diagram(final_merged_file, casrn):
    # Load the Excel file
    simplified_file_path = final_merged_file

    # CAS number provided
    cas_number = casrn

    # Flow Diagram Generation
    df_activity = pd.read_excel(simplified_file_path, sheet_name='Activity')
    df_function = pd.read_excel(simplified_file_path, sheet_name='Function')
    df_naics = pd.read_excel(simplified_file_path, sheet_name='NAICS')
    df_end_of_life = pd.read_excel(simplified_file_path, sheet_name='End-of-Life')
    df_product_use = pd.read_excel(simplified_file_path, sheet_name='Product Use Information')
    df_naics_description = pd.read_excel(simplified_file_path, sheet_name='NAICS Description')

    # Filtered FRS codes based on entries in the Process or Use Operation tab and C-C Product Use Category tab
    df_process = pd.read_excel(simplified_file_path, sheet_name='Process or Use Operation')
    df_cc_product_use_category = pd.read_excel(simplified_file_path, sheet_name='C-C Product Use Category')

    # Combine the FRS codes from both sheets
    valid_frs_codes_process = set(df_process.values.flatten()) - {None}
    valid_frs_codes_cc_product_use = set(df_cc_product_use_category.values.flatten()) - {None}

    # Union of both sets to get all valid FRS codes
    valid_frs_codes = valid_frs_codes_process.union(valid_frs_codes_cc_product_use)


    # Filter the dataframes based on valid FRS codes with debug statements
    df_activity_filtered = filter_frs_codes(df_activity, valid_frs_codes, "Activity")
    df_function_filtered = filter_frs_codes(df_function, valid_frs_codes, "Function")
    df_naics_filtered = filter_frs_codes(df_naics, valid_frs_codes, "NAICS")
    df_end_of_life_filtered = filter_frs_codes(df_end_of_life, valid_frs_codes, "End-of-Life")

    df_product_use_filtered = df_product_use[df_product_use['FRS Code'].isin(valid_frs_codes)].copy()
    df_product_use_filtered['FRS Code'] = df_product_use_filtered['FRS Code'].apply(lambda x: x if x in valid_frs_codes else None)

    # Dynamically create sections and match columns
    function_categories = df_function_filtered.columns
    end_of_life_columns = ['Yes', 'No', 'CBI']
    product_names = df_product_use_filtered['Product Name'].unique()

    # Create a mapping dictionary for end-of-life categories
    end_of_life_mapping = {
        "Yes": "Recycling",
        "No": "Disposal",
        "CBI": "Confidential Business Information"
    }

    # Filter only existing columns in end_of_life_df
    existing_eol_columns = [col for col in end_of_life_columns if col in df_end_of_life_filtered.columns]


    # Combine NAICS sections based on the first 6 digits
    def combine_naics(df):
        combined = {}
        for col in df.columns:
            key = col[:6]
            if key not in combined:
                combined[key] = []
            combined[key].extend(df[col].dropna().tolist())
        combined_df = pd.DataFrame.from_dict(combined, orient='index').transpose()
        return combined_df

    def add_naics_descriptions(df, df_naics_description):
        descriptions = {}
        for naics_code in df.columns:
            if naics_code == "CBI":
                descriptions[naics_code] = "Confidential Business Information"
            else:
                try:
                    description = df_naics_description[df_naics_description['2022 NAICS US   Code'] == int(naics_code)]
                    if not description.empty:
                        descriptions[naics_code] = description['2022 NAICS US Title'].values[0]
                    else:
                        descriptions[naics_code] = ""
                except ValueError:
                    descriptions[naics_code] = ""
        return descriptions

    df_naics_combined = combine_naics(df_naics_filtered)
    naics_descriptions = add_naics_descriptions(df_naics_combined, df_naics_description)


    # Extract unique product names
    unique_product_names = df_product_use_filtered[['Product Name', 'FRS Code']].drop_duplicates()

    # Create custom nodes for sections with split labels
    def create_section_node(label):
        split_label_text = split_label(label, max_length=20)
        # Adjust the two lines below (the ending multiplier numbers, such as * 0.3 --> 0.35 or 0.2) if fontsize="xx" on line 1404 has been modified to avoid text overflow from the white background
        width = max(len(line) for line in split_label_text.split('\n')) * 0.3  # Adjust based on text length
        height = (split_label_text.count('\n') + 1) * 0.75  # Adjust based on line count
        return Custom(split_label_text, "./section_icon.png", fontsize="45", width=str(width), height=str(height), style="filled", fillcolor="white")

    # Define the diagram with spacing and font size adjustments
    graph_attr = {
        "splines": "polyline",  # Use straight lines
        "style": "rounded",  # Rounded corner for edges
        "nodesep": "0.5",  # Adjust node separation
        "ranksep": "8",  # Adjust rank separation
        "fontsize": "60",  # Font size for the diagram
        "fontname": "Arial-BoldMT",  # Bold font for the diagram
        "labeljust": "c",  # Center the label
        "labelloc": "t",  # Place the label at the top
        "label": f"Qualitative Mapping (CAS No. {cas_number})",  # Title with CAS number
        "fontcolor": "black",
        "fontsize": "60",  # Font size of the title
        "rankdir": "TB",  # Sets the direction from top to bottom
    }

    node_attr = {
        "fontsize": "48",  # Font size for FRS nodes
        "fontname": "Arial-BoldMT",  # Font for FRS nodes
        "style": "filled",  # Fill nodes with color
        "fillcolor": "white"  # Set fill color to white
    }

    cluster_attr = {
        "fontsize": "48",  # Font size for cluster labels
        "fontname": "Arial-BoldMT",  # Bold font for cluster labels
        "fontcolor": "black",  # Font color
        "style": "filled",  # Fill clusters with color
        "labeljust": "c",  # Center the label horizontally
        "labelloc": "c",  # Center the label vertically
    }

    section_attr = {
        "fontsize": "48",  # Font size for section labels
        "fontname": "Arial-BoldMT",  # Bold font for section labels
        "fontcolor": "green",  # Font color
        "style": "filled",  # Fill sections with color
        "fillcolor": "white"  # Set fill color to white
    }

    # Function to split label into multiple lines if it exceeds a certain length
    def split_label(label, max_length=20):
        words = label.split()
        lines = []
        current_line = ""
        for word in words:
            if len(current_line) + len(word) + 1 <= max_length:
                current_line += " " + word
            else:
                lines.append(current_line.strip())
                current_line = word
        lines.append(current_line.strip())
        return "\n".join(lines)

    # Create and save the flow diagram using Diagram module
    simplified_diagram_path = f'qualitative_mapping_{cas_number}_Simplified'

    with Diagram("", show=False, direction="TB", filename=simplified_diagram_path, outformat="png", graph_attr=graph_attr, node_attr=node_attr):
        # Manufacturing/Import Cluster
        manufacturing_import_sections = {}
        with Cluster("Manufacturing/Import", graph_attr={**cluster_attr, "bgcolor": "lightblue", "margin": "20,20"}) as manufacturing_import_cluster:
            if "CBI" in df_activity_filtered.columns and df_activity_filtered["CBI"].dropna().any():
                cbi = create_section_node("Confidential Business Information")
                manufacturing_import_sections["CBI"] = cbi
            if "Manufacture" in df_activity_filtered.columns and df_activity_filtered["Manufacture"].dropna().any():
                manufacture = create_section_node("Manufacture")
                manufacturing_import_sections["Manufacture"] = manufacture
            if "Import" in df_activity_filtered.columns and df_activity_filtered["Import"].dropna().any():
                import_ = create_section_node("Import")
                manufacturing_import_sections["Import"] = import_

        # Process or Use Operation Cluster
        process_sections = {}
        with Cluster("Industrial Process or Use Operation", graph_attr={**cluster_attr, "bgcolor": "lightgreen", "margin": "20,20"}) as process_cluster:
            for col in df_process.columns:
                if df_process[col].dropna().any():
                    col_label = split_label(col, max_length=30)
                    process_sections[col] = create_section_node(col_label)

        # C-C Product Use Category Cluster
        cc_product_use_sections = {}
        with Cluster("Consumer and Commercial Product Use Category", graph_attr={**cluster_attr, "style": "filled", "bgcolor": "lightgreen", "margin": "20,20"}) as cc_product_use_category_cluster:
            for col in df_cc_product_use_category.columns:
                if df_cc_product_use_category[col].dropna().any():
                    col_label = split_label(col, max_length=30)
                    cc_product_use_sections[col] = create_section_node(col_label)

        # Function Cluster
        function_sections = {}
        with Cluster("Function", graph_attr={**cluster_attr, "bgcolor": "lightcyan", "margin": "20,20"}) as function_cluster:
            for col in df_function_filtered.columns:
                if df_function_filtered[col].dropna().any():
                    col_label = split_label(col, max_length=30)
                    function_sections[col] = create_section_node(col_label)

        # NAICS Cluster
        naics_sections = {}
        with Cluster("NAICS Codes and Descriptions", graph_attr={**cluster_attr, "bgcolor": "lightgoldenrodyellow", "margin": "20,20"}) as naics_cluster:
            for col in df_naics_combined.columns:
                if df_naics_combined[col].dropna().any():
                    naics_label = split_label(col + "\n" + naics_descriptions.get(col, ""), max_length=30)
                    naics_sections[col] = create_section_node(naics_label)

        # Product Use Information Cluster
        product_use_sections = {}
        with Cluster("Product Use Information", graph_attr={**cluster_attr, "bgcolor": "lightgrey", "margin": "20,20"}) as product_use_cluster:
            if unique_product_names.empty:
                product_label = "No Product Information Found for this Chemical"
                pseudo_node = Custom(product_label, "./icons/blank.png", fontsize="25", style="filled", fillcolor="white")
                product_use_sections[product_label] = pseudo_node
            else:
                for product_name in unique_product_names['Product Name'].unique():
                    product_label = split_label(product_name, max_length=25)
                    product_use_sections[product_name] = create_section_node(product_label)

        # End-of-Life Cluster
        end_of_life_sections = {}
        with Cluster("Industrial End-of-Life", graph_attr={**cluster_attr, "bgcolor": "lightpink", "margin": "20,20"}) as end_of_life_cluster:
            if "CBI" in df_end_of_life_filtered.columns and df_end_of_life_filtered["CBI"].dropna().any():
                eol_cbi = create_section_node("Confidential Business Information")
                end_of_life_sections['Confidential Business Information'] = eol_cbi
            if "Yes" in df_end_of_life_filtered.columns and df_end_of_life_filtered["Yes"].dropna().any():
                eol_yes = create_section_node("Recycling")
                end_of_life_sections['Recycling'] = eol_yes
            if "No" in df_end_of_life_filtered.columns and df_end_of_life_filtered["No"].dropna().any():
                eol_no = create_section_node("Disposal")
                end_of_life_sections['Disposal'] = eol_no
            # Custom edge style
            custom_edge = Edge(color="black", style="bold", penwidth="2.0")
            custom_edge_recycled = Edge(color="darkgreen", style="bold", penwidth="2.0")
            custom_edge_eol = Edge(color="pink", style ="bold", penwidth="2.0")


        # Connect sections based on shared FRS IDs
        def connect_sections(source_df, source_sections, target_df, target_sections, edge_style):
            connected_sections = set()
            for source_col, source_node in source_sections.items():
                for target_col, target_node in target_sections.items():
                    if source_col in source_df.columns and target_col in target_df.columns:
                        shared_frs_ids = set(source_df[source_col].dropna()) & set(target_df[target_col].dropna())
                        if shared_frs_ids and (source_col, target_col) not in connected_sections:
                            source_node >> edge_style >> target_node
                            connected_sections.add((source_col, target_col))

        # Connect Manufacturing/Import sections to Process and C-C Product Use Category sections
        connect_sections(df_activity_filtered, manufacturing_import_sections, df_process, process_sections, custom_edge)
        connect_sections(df_activity_filtered, manufacturing_import_sections, df_cc_product_use_category, cc_product_use_sections, custom_edge)

        # Connect Process sections to Function sections
        connect_sections(df_process, process_sections, df_function_filtered, function_sections, custom_edge)
        # Connect C-C Product Use Category sections to Function sections
        connect_sections(df_cc_product_use_category, cc_product_use_sections, df_function_filtered, function_sections, custom_edge)

        # Connect Function to End-of-Life and NAICS
        connected_sections_function_to_eol = set()
        connected_sections_function_to_naics = set()

        for col in df_function_filtered.columns:
            for frs_id in df_function_filtered[col].dropna():
                for eol_col in existing_eol_columns:
                    mapped_eol_col = end_of_life_mapping[eol_col]  # Map to the correct section key
                    if frs_id in df_end_of_life_filtered[eol_col].dropna().values and (col, mapped_eol_col) not in connected_sections_function_to_eol:
                        if mapped_eol_col in end_of_life_sections:
                            function_sections[col] >> custom_edge_eol >> end_of_life_sections[mapped_eol_col]
                            connected_sections_function_to_eol.add((col, mapped_eol_col))
                for naics_col in df_naics_combined.columns:
                    if frs_id in df_naics_combined[naics_col].dropna().values and (col, naics_col) not in connected_sections_function_to_naics:
                        function_sections[col] >> custom_edge >> naics_sections[naics_col]
                        connected_sections_function_to_naics.add((col, naics_col))

        # Connect NAICS sections to Product Use Information sections based on shared FRS IDs
        connected_sections_naics_to_product = set()
        for col in df_naics_combined.columns:
            for frs_id in df_naics_combined[col].dropna():
                for product_name in unique_product_names['Product Name'].unique():
                    if frs_id in df_product_use_filtered[df_product_use_filtered['Product Name'] == product_name]['FRS Code'].values and (col, product_name) not in connected_sections_naics_to_product:
                        naics_sections[col] >> custom_edge >> product_use_sections[product_name]
                        connected_sections_naics_to_product.add((col, product_name))

        # Connecting End-of-Life sections
        if 'Recycling' in end_of_life_sections:
            eol_recycling = end_of_life_sections['Recycling']
            if 'CBI' in manufacturing_import_sections:
                eol_recycling >> custom_edge_recycled >> manufacturing_import_sections['CBI']
            if 'Manufacture' in manufacturing_import_sections:
                eol_recycling >> custom_edge_recycled >> manufacturing_import_sections['Manufacture']
            if 'Import' in manufacturing_import_sections:
                eol_recycling >> custom_edge_recycled >> manufacturing_import_sections['Import']



    print(f"Simplified diagram file 'qualitative_mapping_{cas_number}_simplified.png' created successfully.")



def print_welcome_message():
    message = """
╔═══════════════════════════════════════════════════════════════════════╗
║                                                                       ║
║       Welcome to the Chemical Condition of Use Locator (ChemCoUL)     ║
║                             Version 1.5.1                             ║
║ This tool accepts a CASRN input from the user(s) to do the following: ║
║                                                                       ║
║    1. Scans through the U.S. Environmental Protection Agency's        ║
║       public databases (i.e., TRI, CDR, and ChemExpo)-version 2020    ║
║                                                                       ║
║    2. Combines the matches into one comprehensive raw data file       ║
║       then summarizes the findings                                    ║
║                                                                       ║
║    3. Reads the summary file and generates a facility summary         ║
║       and qualitative flow diagram                                    ║
║                                                                       ║
╚═══════════════════════════════════════════════════════════════════════╝
    """
    print(message)

def main():
    # Start timing
    start_time = time.time()
    
    # Step 1: Get user input for CASRN or Accession Number
    print_welcome_message()
    user_input = input("Please enter a valid CASRN xx-xx-xx to begin the search: ")

    # Ensure user input is not empty
    if not user_input:
        print("User input is empty. Please enter a valid CASRN in the specified format or try another CASRN.")
        return

    # Define the file path for the qualitative summary file
    qualitative_summary_file = f'qualitative_chemical_flow_mapping_summary_{user_input}.xlsx'
    pucs_naics_file = f'Product_Use_Information_NAICS_Crosswalked_{user_input}.xlsx'

    if os.path.exists(qualitative_summary_file):
        print(f"File '{qualitative_summary_file}' already exists. Skipping steps 1 to 4.")
    else:
        # Step 1: Process CDR and TRI data
        print("Database search initiated. Please be advised that the search time may vary depending on the chemical")
        print("The approximate wait time is 10 - 15 min")
        condition_of_use_file = process_cdr_tri_data(user_input)
        if condition_of_use_file:
            # Check if the Product Use Information_NAICS_Crosswalked file already exists
            if os.path.exists(pucs_naics_file):
                print(f"File '{pucs_naics_file}' already exists. Skipping Step 2.")
                # Proceed to Step 3: Final matching and integration
                final_merged_file = final_matching_and_integration(condition_of_use_file, pucs_naics_file, f'Condition-of-use-TRI1b-CDR-PUCS-NAICS_{user_input}.xlsx')
                if final_merged_file:
                    # Step 4: Generate qualitative summary
                    qualitative_summary_file = generate_qualitative_summary(final_merged_file, user_input)
                    if not qualitative_summary_file:
                        print("Failed to generate qualitative summary.")
                        return
            else:
                # Step 2: Process ChemExpo data
                product_use_info_file = process_chemexpo_data(user_input)
                if product_use_info_file:
                    # Step 2.5: Match with PUCS NAICS Crosswalk
                    pucs_naics_file = match_with_pucs_naics(product_use_info_file, user_input)
                    if pucs_naics_file:
                        # Step 3: Final matching and integration
                        final_merged_file = final_matching_and_integration(condition_of_use_file, pucs_naics_file, f'Condition-of-use-TRI1b-CDR-PUCS-NAICS_{user_input}.xlsx')
                        if final_merged_file:
                            # Step 4: Generate qualitative summary
                            qualitative_summary_file = generate_qualitative_summary(final_merged_file, user_input)
                            if not qualitative_summary_file:
                                print("Failed to generate qualitative summary.")
                                return

    if qualitative_summary_file:
        # Step 5: Generate report and diagram
        generate_report_and_diagram(qualitative_summary_file, user_input)
        # Step 6: Generate simplified flow diagram
        simplified_file_path = qualitative_summary_file
        generate_simplified_diagram(simplified_file_path, user_input)

    # End the timer at the end of your script
    end_time = time.time()

    # Calculate elapsed time
    elapsed_time = end_time - start_time

    # Convert to minutes and seconds
    minutes, seconds = divmod(elapsed_time, 60)

    # Print the elapsed time in a more readable format
    print(f"Total run time: {int(minutes)} minutes and {seconds:.2f} seconds")

if __name__ == "__main__":
    main()
